import openpyxl as xl

file_path = "C:/Automation/Files/DataDriven.xlsx"
sheet_name = "EMPDETAILS"

workbook = xl.load_workbook(file_path)
# sheet = workbook.get_sheet_by_name(sheet_name)
sheet = workbook[sheet_name]

# Number of rows
rowCount = sheet.max_row
print("Total number of rows:", rowCount)

# Number of columns
colCount = sheet.max_column
print("Total number of columns:", colCount)

'''
Printing all values
'''
for r in range(1,rowCount+1):
    for c in range(1,colCount+1):
        print(sheet.cell(row=r,column=c).value,end=" ")
    print()

'''
Printing row by row
'''
# for r in range(1,rowCount+1):
#     emp_id = sheet.cell(row=r+1,column=1).value
#     first_name = sheet.cell(row=r+1,column=2).value
#     print(emp_id," ",first_name)